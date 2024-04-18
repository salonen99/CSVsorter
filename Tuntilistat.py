import os
from pathlib import Path
import pandas as pd
import openpyxl as xl

# Kansiosijainti
__location__ = os.path.realpath(
    os.path.join(os.getcwd(), os.path.dirname(__file__)))

def combine():
    # kansiosijainti
    folder_path = Path.cwd()
    all_files = os.listdir(folder_path)

    # Ei valita muita kuin csv tiedostoja
    csv_files = [f for f in all_files if f.endswith('.csv')]

    # Luodaan lista
    df_list = []

    for csv in csv_files:
        file_path = os.path.join(folder_path, csv)
        try:
            # koitetaan lukea tiedosto UTF-8 enkoodauksella
            df = pd.read_csv(file_path)
            df_list.append(df)
        except UnicodeDecodeError:
            try:
                # jos UTF-8 epäonnistuu koitetaan lukea tiedosto UTF-16 enkoodauksella jossa tab erottaa
                df = pd.read_csv(__location__, sep='\t', encoding='utf-16')
                df_list.append(df)
            except Exception as e:
                print(f"Ei voitu lukea tiedostoa {csv} seuraavan virheen takia: {e}")
        except Exception as e:
            print(f"Ei voitu lukea tiedostoa {csv} seuraavan virheen takia: {e}")

    # liitetään kaikki data yhteen Dataframeen
    try:
        big_df = pd.concat(df_list, ignore_index=True)
    except ValueError:
        # Mikäl kansiossa ei ole csv-tiedostoja, Kerrotaan siitä käyttäjälle ja pysäytetään ohjelma.
        print("Ei csv tiedostoja")
        input()

    # Tallennetaan uutena xlsx tiedostona
    big_df.to_excel(os.path.join(__location__, 'TempExcel.xlsx'), index=False)

def separateToExcel():
    # Ladataan tiedot Väliaikaisesta xlsx tiedostosta ja järjestellään tiedot oikeisiin tiedostoihin oikeille paikoilleen.

    # Ladataan workbook väliaikaisesta xlsx-tiedostosta
    wb = xl.load_workbook(os.path.join(__location__, 'TempExcel.xlsx'))
    ws = wb.active
    # Luodaan lista sarakkeista jotka siirretään
    columns = ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]
    columns2 =["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
    row = 1
    row2 = 10
    for cell in ws["C"]:
        row += 1
        row2 += 1
        if cell.value != None and ws["C" + str(row)].value != None:
            name = ws["C" + str(row)].value
            # Korjataan nimi
            name = name.replace("-", " ")
            name = name.title()
            try:
                # jos työntekijälle löytyy jo xlsx tiedosto, etsitään seuraava tyhjä solu.
                wb2 = xl.load_workbook(os.path.join(__location__, name + ' Tuntilista.xlsx'))
                ws2 = wb2.active
                for cell in ws2["C"]:
                    if ws2["c" + str(row2)].value != None:
                        row2 += 1
            except FileNotFoundError:
                # jos työntekijälle ei löäydy vielä xlsx tiedostoa, luodaan sellainen.
                try:
                    wb2 = xl.load_workbook(os.path.join(__location__, 'Tuntilista_Empty.xlsx'))
                    ws2 = wb2.active
                    ws2["M5"].value = name
                except FileNotFoundError:
                    # Jos tyhjä pohjatiedosto puutuu keskeytetään ohjelma
                    print("Tiedosto Tuntilista_Empty.xlsx puuttuu.\nKeskeytetään toimenpide")
                    input()
            # Järjestellään arvot oikeille paikoilleen
            for column in columns:
                if column == "N":
                    ws2["M"+str(row2)].value = ws[column + str(row)].value
                if column == "O":
                    ws2["N"+str(row2)].value = ws[column + str(row)].value
                col = columns.index(column)
                try:
                    col2 = columns2[col]
                    ws2[col2 +str(row2)].value = ws[column + str(row)].value
                except IndexError:
                    pass
        try:
            # Tallennetaan tiedostot
            wb2.save(os.path.join(__location__, name + ' Tuntilista.xlsx'))
        except FileNotFoundError:
            pass

def deletefiles():
    # kansiosijainti
    folder_path = Path.cwd()
    all_files = os.listdir(folder_path)

    # Ei valita muita kuin csv tiedostoja
    csv_files = [f for f in all_files if f.endswith('.csv')]
    for csv in csv_files:
        file_path = os.path.join(folder_path, csv)
        os.remove(file_path)
    
    # poistetaan väliaikainen xlsx-tiedosto
    os.remove(os.path.join(__location__, 'TempExcel.xlsx'))

def main():
    combine()
    separateToExcel()
    deletefiles()

main()